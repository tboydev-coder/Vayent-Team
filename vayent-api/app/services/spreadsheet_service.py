"""Spreadsheet source ingestion, profiling, and deterministic BI helpers."""
from __future__ import annotations

import csv
import io
import ipaddress
import logging
import math
import re
import socket
import uuid
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlencode, urlparse, urlunparse

import httpx
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import get_settings
from app.models import SpreadsheetSource, SpreadsheetSourceKind

logger = logging.getLogger(__name__)


class SpreadsheetValidationError(ValueError):
    """Raised when a spreadsheet upload or link is not supported."""


@dataclass
class ProcessedSpreadsheet:
    """Parsed spreadsheet metadata and bounded row payload."""

    file_type: str
    raw_schema_metadata: dict[str, Any]
    dataset_payload: dict[str, Any]
    analysis_metadata: dict[str, Any]


class SpreadsheetService:
    """Manage spreadsheet sources as first-class Vayent data sources."""

    NUMERIC_COLUMN_TYPES = {"number", "currency", "percentage"}
    SEGMENT_ROLES = {
        "category",
        "customer",
        "product",
        "team",
        "geography",
        "status",
        "vendor",
        "employee",
        "channel",
    }
    ROLE_LABELS = {
        "date": "Date",
        "revenue": "Revenue",
        "cost": "Cost",
        "profit": "Profit",
        "percentage": "Rate",
        "quantity": "Quantity",
        "customer": "Customer",
        "product": "Product",
        "team": "Team",
        "geography": "Geography",
        "status": "Operational Status",
        "vendor": "Vendor",
        "employee": "Employee",
        "channel": "Channel",
        "identifier": "Identifier",
        "metric": "Metric",
        "category": "Category",
        "text": "Text",
    }
    CURRENCY_SYMBOL_PATTERN = re.compile(r"[$\u00a3\u20ac\u00a5\u20b9]")

    ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".csv"}
    EXTENSION_TO_FILE_TYPE = {
        ".xlsx": "xlsx",
        ".xls": "xls",
        ".csv": "csv",
    }
    SPREADSHEET_CONTENT_TYPES = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "text/csv",
        "application/csv",
        "text/plain",
        "application/octet-stream",
        "binary/octet-stream",
    }

    def __init__(self) -> None:
        self.settings = get_settings()

    @staticmethod
    def _json_safe(value: Any) -> Any:
        try:
            return json_loads(json_dumps(value))
        except Exception:
            return str(value)

    @staticmethod
    def _clean_name(value: str, fallback: str = "Spreadsheet") -> str:
        cleaned = re.sub(r"\s+", " ", value.strip())
        return cleaned[:255] or fallback

    @staticmethod
    def _humanize_identifier(value: str) -> str:
        cleaned = re.sub(r"[_-]+", " ", value.strip())
        cleaned = re.sub(r"([a-z0-9])([A-Z])", r"\1 \2", cleaned)
        cleaned = re.sub(r"\s+", " ", cleaned).strip()
        return cleaned.title() if cleaned else "Value"

    @classmethod
    def _extension_from_name(cls, filename: str | None) -> str:
        extension = Path(filename or "").suffix.lower()
        if extension not in cls.ALLOWED_EXTENSIONS:
            allowed = ", ".join(sorted(cls.ALLOWED_EXTENSIONS))
            raise SpreadsheetValidationError(
                f"Unsupported file type. Upload a spreadsheet file ({allowed})."
            )
        return extension

    @staticmethod
    def _normalize_header(value: Any, index: int, seen: set[str]) -> str:
        raw = "" if value is None else str(value).strip()
        if not raw:
            raw = f"column_{index + 1}"
        normalized = re.sub(r"\s+", " ", raw)[:120]
        base = normalized
        counter = 2
        while normalized.lower() in seen:
            normalized = f"{base} {counter}"
            counter += 1
        seen.add(normalized.lower())
        return normalized

    @staticmethod
    def _is_blank_row(values: list[Any]) -> bool:
        return all(value in (None, "") for value in values)

    @staticmethod
    def _coerce_cell(value: Any) -> Any:
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        if isinstance(value, str):
            return re.sub(r"\s+", " ", value).strip()
        return value

    def _validate_size(self, content: bytes) -> None:
        max_bytes = max(self.settings.spreadsheet_max_file_size_mb, 1) * 1024 * 1024
        if len(content) > max_bytes:
            raise SpreadsheetValidationError(
                f"Spreadsheet is too large. Maximum size is {self.settings.spreadsheet_max_file_size_mb} MB."
            )
        if not content:
            raise SpreadsheetValidationError("Spreadsheet file is empty.")

    def _validate_file_signature(self, *, extension: str, content: bytes) -> None:
        if extension == ".xlsx" and not content.startswith(b"PK"):
            raise SpreadsheetValidationError("The uploaded .xlsx file is not a valid Excel workbook.")
        if extension == ".xls" and not content.startswith(b"\xd0\xcf\x11\xe0"):
            raise SpreadsheetValidationError("The uploaded .xls file is not a valid Excel workbook.")

    def validate_upload(self, *, filename: str | None, content_type: str | None, content: bytes) -> str:
        extension = self._extension_from_name(filename)
        self._validate_size(content)
        if content_type:
            normalized_content_type = content_type.split(";")[0].strip().lower()
            if (
                normalized_content_type
                and normalized_content_type not in self.SPREADSHEET_CONTENT_TYPES
            ):
                raise SpreadsheetValidationError(
                    "Unsupported upload format. Use .xlsx, .xls, or .csv."
                )
        self._validate_file_signature(extension=extension, content=content)
        return self.EXTENSION_TO_FILE_TYPE[extension]

    def _decode_csv(self, content: bytes) -> str:
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                return content.decode(encoding)
            except UnicodeDecodeError:
                continue
        raise SpreadsheetValidationError("CSV file could not be decoded.")

    def _parse_csv(self, content: bytes) -> dict[str, Any]:
        text = self._decode_csv(content)
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(io.StringIO(text), dialect)
        rows = [row for row in reader if not self._is_blank_row(row)]
        if not rows:
            raise SpreadsheetValidationError("Spreadsheet does not contain any rows.")
        return self._rows_to_table("Sheet1", rows)

    def _parse_xlsx(self, content: bytes) -> list[dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise SpreadsheetValidationError(
                "Excel .xlsx support is not installed on this server."
            ) from exc

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        tables: list[dict[str, Any]] = []
        for worksheet in workbook.worksheets:
            rows = [
                list(row)
                for row in worksheet.iter_rows(values_only=True)
                if not self._is_blank_row(list(row))
            ]
            if rows:
                tables.append(self._rows_to_table(worksheet.title, rows))
        if not tables:
            raise SpreadsheetValidationError("Spreadsheet does not contain any rows.")
        return tables

    def _parse_xls(self, content: bytes) -> list[dict[str, Any]]:
        try:
            import xlrd
        except ImportError as exc:
            raise SpreadsheetValidationError(
                "Excel .xls support is not installed on this server."
            ) from exc

        workbook = xlrd.open_workbook(file_contents=content)
        tables: list[dict[str, Any]] = []
        for sheet in workbook.sheets():
            rows = [
                [sheet.cell_value(row_index, column_index) for column_index in range(sheet.ncols)]
                for row_index in range(sheet.nrows)
            ]
            rows = [row for row in rows if not self._is_blank_row(row)]
            if rows:
                tables.append(self._rows_to_table(sheet.name, rows))
        if not tables:
            raise SpreadsheetValidationError("Spreadsheet does not contain any rows.")
        return tables

    def _rows_to_table(self, sheet_name: str, raw_rows: list[list[Any]]) -> dict[str, Any]:
        if not raw_rows:
            raise SpreadsheetValidationError("Spreadsheet does not contain any rows.")

        seen_headers: set[str] = set()
        headers = [
            self._normalize_header(value, index, seen_headers)
            for index, value in enumerate(raw_rows[0])
        ]
        data_rows = raw_rows[1:]
        max_rows = max(self.settings.spreadsheet_max_rows, 1)
        if len(data_rows) > max_rows:
            data_rows = data_rows[:max_rows]

        records: list[dict[str, Any]] = []
        for raw_row in data_rows:
            record: dict[str, Any] = {}
            for index, header in enumerate(headers):
                value = raw_row[index] if index < len(raw_row) else None
                record[header] = self._coerce_cell(value)
            if any(value not in (None, "") for value in record.values()):
                records.append(record)

        if not headers:
            raise SpreadsheetValidationError("Spreadsheet does not contain columns.")

        return {
            "name": self._clean_name(sheet_name, "Sheet"),
            "row_count": len(records),
            "columns": self._profile_columns(headers, records),
            "rows": records,
            "preview_rows": records[: max(self.settings.spreadsheet_preview_rows, 1)],
            "truncated": len(raw_rows[1:]) > max_rows,
        }

    def _profile_columns(
        self,
        headers: list[str],
        rows: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        columns: list[dict[str, Any]] = []
        for header in headers:
            values = [row.get(header) for row in rows]
            present_values = [value for value in values if value not in (None, "")]
            data_type = self._infer_column_type(present_values)
            categories = []
            if data_type in {"category", "text"}:
                counts: dict[str, int] = {}
                for value in present_values:
                    text = str(value)
                    counts[text] = counts.get(text, 0) + 1
                categories = [
                    {"value": key, "count": count}
                    for key, count in sorted(
                        counts.items(),
                        key=lambda item: item[1],
                        reverse=True,
                    )[:8]
                ]

            numeric_values = [
                self._to_number(value)
                for value in present_values
                if self._to_number(value) is not None
            ]
            date_values = [
                parsed
                for parsed in (self._to_datetime(value) for value in present_values)
                if parsed is not None
            ]
            role = self._infer_column_role(
                header=header,
                data_type=data_type,
                values=present_values,
                numeric_values=numeric_values,
            )
            business_label = self.ROLE_LABELS.get(role) or self._humanize_identifier(header)

            columns.append(
                {
                    "name": header,
                    "label": self._humanize_identifier(header),
                    "business_label": business_label,
                    "type": data_type,
                    "semantic_role": role,
                    "nullable": len(present_values) < len(values),
                    "non_empty_count": len(present_values),
                    "unique_count": len({str(value) for value in present_values}),
                    "categories": categories,
                    "min": min(numeric_values) if numeric_values else None,
                    "max": max(numeric_values) if numeric_values else None,
                    "sum": sum(numeric_values) if numeric_values else None,
                    "average": (
                        sum(numeric_values) / len(numeric_values)
                        if numeric_values
                        else None
                    ),
                    "date_min": min(date_values).date().isoformat() if date_values else None,
                    "date_max": max(date_values).date().isoformat() if date_values else None,
                    "examples": [self._json_safe(value) for value in present_values[:3]],
                }
            )
        return columns

    @staticmethod
    def _to_number(value: Any) -> float | None:
        if isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            numeric = float(value)
            return numeric if math.isfinite(numeric) else None
        if isinstance(value, str):
            normalized = value.strip()
            if not normalized:
                return None
            is_negative = normalized.startswith("(") and normalized.endswith(")")
            normalized = normalized.strip("()")
            normalized = re.sub(r"[$\u00a3\u20ac\u00a5\u20b9,\s%]", "", normalized)
            if not normalized:
                return None
            try:
                numeric = float(normalized)
                if is_negative:
                    numeric *= -1
                return numeric if math.isfinite(numeric) else None
            except ValueError:
                return None
        return None

    @staticmethod
    def _to_datetime(value: Any) -> datetime | None:
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime.combine(value, datetime.min.time())
        if not isinstance(value, str):
            return None

        candidate = value.strip()
        if not candidate:
            return None

        normalized = candidate.replace("Z", "+00:00")
        try:
            return datetime.fromisoformat(normalized)
        except ValueError:
            pass

        for date_format in (
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%m/%d/%Y",
            "%d/%m/%Y",
            "%m-%d-%Y",
            "%d-%m-%Y",
            "%b %Y",
            "%B %Y",
            "%Y-%m",
            "%Y/%m",
            "%Y",
        ):
            try:
                return datetime.strptime(candidate, date_format)
            except ValueError:
                continue
        return None

    @staticmethod
    def _looks_like_date(value: Any) -> bool:
        return SpreadsheetService._to_datetime(value) is not None

    def _infer_column_type(self, values: list[Any]) -> str:
        if not values:
            return "empty"
        numeric_count = sum(1 for value in values if self._to_number(value) is not None)
        date_count = sum(1 for value in values if self._looks_like_date(value))
        bool_count = sum(1 for value in values if isinstance(value, bool))
        threshold = max(1, int(len(values) * 0.7))
        if bool_count >= threshold:
            return "boolean"
        if numeric_count >= threshold:
            if self._has_percentage_signal("", values):
                return "percentage"
            if self._has_currency_signal("", values):
                return "currency"
            return "number"
        if date_count >= threshold:
            return "date"
        unique_count = len({str(value) for value in values})
        if unique_count <= max(12, len(values) * 0.3):
            return "category"
        return "text"

    @staticmethod
    def _normalized_identifier(value: str) -> str:
        return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()

    def _has_currency_signal(self, header: str, values: list[Any]) -> bool:
        normalized_header = self._normalized_identifier(header)
        currency_terms = {
            "amount",
            "arpu",
            "arr",
            "balance",
            "budget",
            "cost",
            "expense",
            "fee",
            "gmv",
            "invoice",
            "mrr",
            "payment",
            "price",
            "profit",
            "revenue",
            "sales",
            "spend",
            "total",
        }
        if any(term in normalized_header.split() for term in currency_terms):
            return True
        return any(
            isinstance(value, str) and self.CURRENCY_SYMBOL_PATTERN.search(value)
            for value in values[:50]
        )

    def _has_percentage_signal(self, header: str, values: list[Any]) -> bool:
        normalized_header = self._normalized_identifier(header)
        percentage_terms = {
            "conversion",
            "margin",
            "pct",
            "percent",
            "percentage",
            "rate",
            "ratio",
            "retention",
            "roi",
            "share",
            "yield",
        }
        if any(term in normalized_header.split() for term in percentage_terms):
            return True
        return any(isinstance(value, str) and "%" in value for value in values[:50])

    def _infer_column_role(
        self,
        *,
        header: str,
        data_type: str,
        values: list[Any],
        numeric_values: list[float],
    ) -> str:
        normalized = self._normalized_identifier(header)
        words = set(normalized.split())

        def has_any(*terms: str) -> bool:
            return any(term in normalized or term in words for term in terms)

        if data_type == "date" or has_any(
            "date",
            "month",
            "week",
            "year",
            "period",
            "created",
            "updated",
            "ordered",
            "paid",
            "shipped",
            "closed",
        ):
            return "date"

        if has_any("customer", "client", "buyer", "account", "subscriber", "member"):
            return "customer"
        if has_any("team", "club", "organisation", "organization", "group"):
            return "team"
        if has_any("product", "sku", "item", "service", "plan", "package", "category"):
            return "product" if has_any("product", "sku", "item", "service", "plan", "package") else "category"
        if has_any("country", "region", "city", "state", "market", "territory", "location", "zip", "postcode"):
            return "geography"
        if has_any("status", "stage", "state", "outcome", "priority", "queue"):
            return "status"
        if has_any("vendor", "supplier", "partner", "merchant"):
            return "vendor"
        if has_any("employee", "agent", "owner", "rep", "staff"):
            return "employee"
        if has_any("channel", "campaign", "source", "medium"):
            return "channel"
        if has_any(" id", "_id", "identifier") or normalized.endswith(" id") or normalized == "id":
            return "identifier"

        if data_type in self.NUMERIC_COLUMN_TYPES:
            if data_type == "percentage" or self._has_percentage_signal(header, values):
                return "percentage"
            if has_any("profit", "margin", "gross profit", "net profit"):
                return "profit"
            if has_any("cost", "expense", "spend", "cogs", "refund", "discount"):
                return "cost"
            if has_any("revenue", "sales", "amount", "gmv", "mrr", "arr", "income", "booking", "payment", "invoice", "total"):
                return "revenue"
            if has_any("quantity", "qty", "unit", "units", "volume", "count", "orders", "tickets"):
                return "quantity"
            if self._has_currency_signal(header, values):
                return "revenue"
            return "metric"

        if data_type == "category":
            return "category"
        if data_type == "text":
            unique_count = len({str(value) for value in values})
            if values and unique_count <= max(20, len(values) * 0.35):
                return "category"
        return "text"

    @classmethod
    def _is_numeric_column(cls, column: dict[str, Any]) -> bool:
        return (
            column.get("type") in cls.NUMERIC_COLUMN_TYPES
            and column.get("semantic_role") != "identifier"
        )

    @classmethod
    def _is_segment_column(cls, column: dict[str, Any]) -> bool:
        role = str(column.get("semantic_role") or "")
        column_type = str(column.get("type") or "")
        unique_count = int(column.get("unique_count") or 0)
        non_empty_count = int(column.get("non_empty_count") or 0)
        if role in cls.SEGMENT_ROLES:
            return True
        if column_type == "category":
            return True
        return (
            column_type == "text"
            and non_empty_count > 0
            and unique_count <= max(25, non_empty_count * 0.4)
        )

    @staticmethod
    def _metric_priority(column: dict[str, Any]) -> int:
        role = str(column.get("semantic_role") or "")
        return {
            "revenue": 100,
            "profit": 92,
            "cost": 84,
            "quantity": 76,
            "percentage": 68,
            "metric": 60,
        }.get(role, 50)

    def _select_primary_metric(self, columns: list[dict[str, Any]]) -> dict[str, Any] | None:
        numeric_columns = [column for column in columns if self._is_numeric_column(column)]
        if not numeric_columns:
            return None
        return sorted(
            numeric_columns,
            key=lambda column: (
                self._metric_priority(column),
                column.get("non_empty_count") or 0,
                abs(float(column.get("sum") or 0)),
            ),
            reverse=True,
        )[0]

    @staticmethod
    def _select_primary_date(columns: list[dict[str, Any]]) -> dict[str, Any] | None:
        date_columns = [
            column
            for column in columns
            if column.get("type") == "date" or column.get("semantic_role") == "date"
        ]
        if not date_columns:
            return None
        return sorted(
            date_columns,
            key=lambda column: column.get("non_empty_count") or 0,
            reverse=True,
        )[0]

    def _select_segment_columns(
        self,
        columns: list[dict[str, Any]],
        *,
        limit: int = 4,
    ) -> list[dict[str, Any]]:
        role_priority = {
            "product": 90,
            "customer": 84,
            "team": 82,
            "geography": 78,
            "channel": 72,
            "status": 66,
            "vendor": 62,
            "employee": 58,
            "category": 54,
        }
        candidates = [
            column
            for column in columns
            if self._is_segment_column(column)
            and column.get("semantic_role") != "identifier"
        ]
        return sorted(
            candidates,
            key=lambda column: (
                role_priority.get(str(column.get("semantic_role") or ""), 50),
                column.get("non_empty_count") or 0,
                -(column.get("unique_count") or 0),
            ),
            reverse=True,
        )[:limit]

    @staticmethod
    def _column_name(column: dict[str, Any] | None) -> str:
        return str((column or {}).get("name") or "")

    def _metric_label(self, column: dict[str, Any] | None) -> str:
        if not column:
            return "metric"
        role = str(column.get("semantic_role") or "")
        if role in {"revenue", "cost", "profit", "percentage", "quantity"}:
            return self.ROLE_LABELS.get(role, self._humanize_identifier(self._column_name(column)))
        return self._humanize_identifier(self._column_name(column))

    def _segment_label(self, column: dict[str, Any] | None) -> str:
        if not column:
            return "segment"
        role = str(column.get("semantic_role") or "")
        if role in self.ROLE_LABELS:
            return self.ROLE_LABELS[role].lower()
        return self._humanize_identifier(self._column_name(column)).lower()

    def _format_metric_value(self, value: float | int | None, column: dict[str, Any] | None) -> str:
        if value is None:
            return "n/a"
        numeric = float(value)
        column_type = str((column or {}).get("type") or "")
        role = str((column or {}).get("semantic_role") or "")
        max_value = (column or {}).get("max")

        if column_type == "percentage" or role == "percentage":
            max_abs = abs(float(max_value or 0))
            percent_value = numeric * 100 if max_abs <= 1.5 and abs(numeric) <= 1.5 else numeric
            return f"{percent_value:,.1f}%"

        if column_type == "currency" or role in {"revenue", "cost", "profit"}:
            decimals = 0 if abs(numeric) >= 100 else 2
            return f"${numeric:,.{decimals}f}"

        decimals = 0 if abs(numeric) >= 100 or float(numeric).is_integer() else 1
        return f"{numeric:,.{decimals}f}"

    @staticmethod
    def _format_percent_delta(previous: float, latest: float) -> str:
        if previous == 0:
            return "from zero"
        delta = ((latest - previous) / abs(previous)) * 100
        direction = "increased" if delta >= 0 else "decreased"
        return f"{direction} {abs(delta):.1f}%"

    def _date_bucket(self, value: Any) -> tuple[str, datetime] | None:
        parsed = self._to_datetime(value)
        if parsed is None:
            return None
        return parsed.strftime("%Y-%m"), datetime(parsed.year, parsed.month, 1)

    def _aggregate_by_period(
        self,
        rows: list[dict[str, Any]],
        date_key: str,
        numeric_key: str,
    ) -> list[dict[str, Any]]:
        buckets: dict[str, dict[str, Any]] = {}
        for row in rows:
            bucket = self._date_bucket(row.get(date_key))
            value = self._to_number(row.get(numeric_key))
            if bucket is None or value is None:
                continue
            label, sort_value = bucket
            if label not in buckets:
                buckets[label] = {
                    "period": label,
                    "value": 0.0,
                    "count": 0,
                    "_sort": sort_value,
                }
            buckets[label]["value"] += value
            buckets[label]["count"] += 1
        return [
            {key: value for key, value in item.items() if key != "_sort"}
            for item in sorted(buckets.values(), key=lambda item: item["_sort"])
        ]

    def _count_categories(
        self,
        rows: list[dict[str, Any]],
        category_key: str,
    ) -> list[dict[str, Any]]:
        counts: Counter[str] = Counter()
        for row in rows:
            value = row.get(category_key)
            if value in (None, ""):
                continue
            counts[str(value)] += 1

        total = sum(counts.values())
        if total == 0:
            return []
        return [
            {
                "segment": segment,
                "count": count,
                "share": (count / total) * 100,
            }
            for segment, count in counts.most_common()
        ]

    def _row_label(
        self,
        row: dict[str, Any],
        *,
        date_key: str | None,
        segment_keys: list[str],
    ) -> str:
        label_parts = []
        if date_key and row.get(date_key) not in (None, ""):
            label_parts.append(str(row.get(date_key)))
        for key in segment_keys:
            if row.get(key) not in (None, ""):
                label_parts.append(str(row.get(key)))
            if len(label_parts) >= 2:
                break
        return " / ".join(label_parts) or "Row"

    def _detect_numeric_anomalies(
        self,
        rows: list[dict[str, Any]],
        metric: dict[str, Any],
        *,
        date_key: str | None,
        segment_keys: list[str],
    ) -> list[dict[str, Any]]:
        metric_key = self._column_name(metric)
        values = []
        for row in rows:
            numeric = self._to_number(row.get(metric_key))
            if numeric is None:
                continue
            values.append((row, numeric))

        if len(values) < 4:
            return []

        numeric_values = [value for _, value in values]
        mean = sum(numeric_values) / len(numeric_values)
        variance = sum((value - mean) ** 2 for value in numeric_values) / len(numeric_values)
        stdev = math.sqrt(variance)
        if stdev == 0:
            return []

        anomalies = []
        for row, value in values:
            score = abs(value - mean) / stdev
            if score < 2:
                continue
            anomalies.append(
                {
                    "label": self._row_label(
                        row,
                        date_key=date_key,
                        segment_keys=segment_keys,
                    ),
                    "value": value,
                    "average": mean,
                    "z_score": score,
                }
            )
        return sorted(anomalies, key=lambda item: item["z_score"], reverse=True)[:5]

    def _build_quality_checks(
        self,
        *,
        table: dict[str, Any],
        columns: list[dict[str, Any]],
        rows: list[dict[str, Any]],
    ) -> dict[str, Any]:
        row_count = len(rows)
        missing_values = []
        empty_columns = []
        invalid_dates = []

        for column in columns:
            name = self._column_name(column)
            non_empty_count = int(column.get("non_empty_count") or 0)
            missing_count = max(row_count - non_empty_count, 0)
            if missing_count:
                missing_values.append(
                    {
                        "column": name,
                        "missing_count": missing_count,
                        "missing_percent": (missing_count / row_count * 100) if row_count else 0,
                    }
                )
            if non_empty_count == 0:
                empty_columns.append(name)
            if column.get("semantic_role") == "date" or column.get("type") == "date":
                invalid_count = sum(
                    1
                    for row in rows
                    if row.get(name) not in (None, "")
                    and self._to_datetime(row.get(name)) is None
                )
                if invalid_count:
                    invalid_dates.append(
                        {
                            "column": name,
                            "invalid_count": invalid_count,
                        }
                    )

        fingerprints = Counter(
            json_dumps(
                {
                    key: row.get(key)
                    for key in sorted(row.keys())
                    if row.get(key) not in (None, "")
                }
            )
            for row in rows
        )
        duplicate_records = sum(count - 1 for count in fingerprints.values() if count > 1)

        return {
            "table": table.get("name"),
            "row_count": row_count,
            "column_count": len(columns),
            "missing_values": sorted(
                missing_values,
                key=lambda item: item["missing_count"],
                reverse=True,
            )[:10],
            "empty_columns": empty_columns,
            "duplicate_records": duplicate_records,
            "invalid_dates": invalid_dates,
        }

    def _build_suggested_questions(
        self,
        *,
        columns: list[dict[str, Any]],
        table: dict[str, Any],
    ) -> list[str]:
        questions = []
        primary_metric = self._select_primary_metric(columns)
        primary_date = self._select_primary_date(columns)
        segments = self._select_segment_columns(columns)
        entity = self._select_entity_column(columns)
        status = next(
            (column for column in columns if column.get("semantic_role") == "status"),
            None,
        )

        entity_label = self._segment_label(entity).lower() if entity else "records"
        if entity:
            questions.append(f"How many {entity_label}s are registered?")
            questions.append(f"List 10 {entity_label}s.")
        else:
            questions.append("How many records are in this file?")
            questions.append("Show me the first 20 records.")

        if primary_date and entity:
            questions.append(f"Which {entity_label}s registered most recently?")
        elif primary_date:
            questions.append("What is the trend over time?")

        if status:
            status_label = self._humanize_identifier(self._column_name(status)).lower()
            questions.append(f"How many records are in each {status_label}?")
            for category in status.get("categories", [])[:2]:
                value = category.get("value")
                if value:
                    questions.append(f"How many {str(value).lower()} {entity_label}s exist?")

        if primary_metric and segments:
            questions.append(
                f"Which {self._segment_label(segments[0])} has the highest {self._metric_label(primary_metric).lower()}?"
            )
        questions.append("Show duplicates.")

        deduped: list[str] = []
        for question in questions:
            if question not in deduped:
                deduped.append(question)
        return deduped[:8]

    def _select_entity_column(self, columns: list[dict[str, Any]]) -> dict[str, Any] | None:
        entity_roles = {"team", "customer", "product", "vendor", "employee"}
        entities = [
            column
            for column in columns
            if column.get("semantic_role") in entity_roles
        ]
        if entities:
            return sorted(
                entities,
                key=lambda column: (
                    {"team": 90, "customer": 86, "product": 82}.get(
                        str(column.get("semantic_role") or ""),
                        70,
                    ),
                    column.get("non_empty_count") or 0,
                ),
                reverse=True,
            )[0]
        text_columns = [
            column
            for column in columns
            if column.get("type") == "text"
            and column.get("semantic_role") not in {"identifier", "date"}
        ]
        if not text_columns:
            return None
        return sorted(
            text_columns,
            key=lambda column: (
                column.get("unique_count") or 0,
                column.get("non_empty_count") or 0,
            ),
            reverse=True,
        )[0]

    def _build_analysis(self, tables: list[dict[str, Any]]) -> dict[str, Any]:
        table_summaries: list[dict[str, Any]] = []
        dataset_understanding: list[dict[str, Any]] = []
        insights: list[dict[str, Any]] = []
        recommendations: list[dict[str, Any]] = []
        dashboard_cards: list[dict[str, Any]] = []
        kpis: list[dict[str, Any]] = []
        risks: list[dict[str, Any]] = []
        opportunities: list[dict[str, Any]] = []
        quality_checks: list[dict[str, Any]] = []
        suggested_questions: list[str] = []
        seen_insights: set[tuple[str, str]] = set()
        seen_recommendations: set[tuple[str, str]] = set()

        def add_insight(
            *,
            title: str,
            body: str,
            tone: str = "neutral",
            priority_score: int = 50,
            evidence: dict[str, Any] | None = None,
        ) -> None:
            title = title.strip()
            body = body.strip()
            if not title or not body:
                return
            key = (title.lower(), body.lower())
            if key in seen_insights:
                return
            seen_insights.add(key)
            insights.append(
                {
                    "title": title,
                    "body": body,
                    "tone": tone,
                    "priority_score": priority_score,
                    "evidence": evidence or {},
                }
            )

        def add_recommendation(
            *,
            title: str,
            body: str,
            priority: str = "Medium",
            priority_score: int = 50,
            evidence: dict[str, Any] | None = None,
        ) -> None:
            title = title.strip()
            body = body.strip()
            if not title or not body:
                return
            key = (title.lower(), body.lower())
            if key in seen_recommendations:
                return
            seen_recommendations.add(key)
            recommendations.append(
                {
                    "title": title,
                    "body": body,
                    "priority": priority,
                    "priority_score": priority_score,
                    "evidence": evidence or {},
                }
            )

        for table in tables:
            columns = table.get("columns", [])
            rows = table.get("rows", [])
            numeric_columns = [col for col in columns if self._is_numeric_column(col)]
            segment_columns = self._select_segment_columns(columns)
            date_columns = [
                col
                for col in columns
                if col.get("type") == "date" or col.get("semantic_role") == "date"
            ]
            primary_metric = self._select_primary_metric(columns)
            primary_date = self._select_primary_date(columns)
            primary_metric_key = self._column_name(primary_metric)
            primary_date_key = self._column_name(primary_date)
            segment_keys = [self._column_name(column) for column in segment_columns]

            table_summaries.append(
                {
                    "name": table.get("name"),
                    "row_count": table.get("row_count", 0),
                    "column_count": len(columns),
                    "numeric_columns": [col.get("name") for col in numeric_columns],
                    "category_columns": [col.get("name") for col in segment_columns],
                    "date_columns": [col.get("name") for col in date_columns],
                    "primary_metric": primary_metric_key or None,
                    "primary_date": primary_date_key or None,
                }
            )
            quality_checks.append(
                self._build_quality_checks(
                    table=table,
                    columns=columns,
                    rows=rows,
                )
            )
            for question in self._build_suggested_questions(
                columns=columns,
                table=table,
            ):
                if question not in suggested_questions:
                    suggested_questions.append(question)
            dataset_understanding.append(
                {
                    "table": table.get("name"),
                    "row_count": table.get("row_count", 0),
                    "primary_metric": {
                        "column": primary_metric_key,
                        "label": self._metric_label(primary_metric),
                        "role": (primary_metric or {}).get("semantic_role"),
                    }
                    if primary_metric
                    else None,
                    "date_column": primary_date_key or None,
                    "segments": [
                        {
                            "column": self._column_name(column),
                            "role": column.get("semantic_role"),
                            "label": self._segment_label(column),
                        }
                        for column in segment_columns
                    ],
                    "columns": [
                        {
                            "name": column.get("name"),
                            "type": column.get("type"),
                            "semantic_role": column.get("semantic_role"),
                            "business_label": column.get("business_label"),
                        }
                        for column in columns
                    ],
                }
            )

            if rows:
                dashboard_cards.append(
                    {
                        "title": f"{table.get('name', 'Spreadsheet')} records",
                        "description": "How much spreadsheet evidence Vayent analyzed.",
                        "visualization": "kpi",
                        "value": table.get("row_count", 0),
                        "rows": [{"records": table.get("row_count", 0)}],
                        "explanation": f"{table.get('row_count', 0)} rows are available for analysis.",
                        "interpretation": "This is the evidence base behind the spreadsheet insights.",
                        "recommended_action": "Use the ranked findings below to decide where to focus first.",
                        "status": "success",
                    }
                )

            if primary_metric and isinstance(primary_metric.get("sum"), (int, float)):
                metric_label = self._metric_label(primary_metric)
                total = float(primary_metric.get("sum") or 0)
                kpi = {
                    "table": table.get("name"),
                    "metric": primary_metric_key,
                    "label": metric_label,
                    "value": total,
                    "formatted_value": self._format_metric_value(total, primary_metric),
                    "average": primary_metric.get("average"),
                    "role": primary_metric.get("semantic_role"),
                }
                kpis.append(kpi)
                add_insight(
                    title=f"{metric_label} is the lead KPI",
                    body=(
                        f"{metric_label} totals {self._format_metric_value(total, primary_metric)} "
                        f"across {table.get('row_count', len(rows)):,} spreadsheet records."
                    ),
                    tone="positive",
                    priority_score=76,
                    evidence=kpi,
                )
                dashboard_cards.append(
                    {
                        "title": metric_label,
                        "description": f"Total {metric_label.lower()} from the spreadsheet.",
                        "visualization": "kpi",
                        "value": total,
                        "rows": [{primary_metric_key: total}],
                        "explanation": f"{metric_label} is the strongest numeric business signal Vayent found.",
                        "interpretation": f"The current total is {self._format_metric_value(total, primary_metric)}.",
                        "recommended_action": (
                            f"Review {metric_label.lower()} by the strongest segment to see where performance is concentrated."
                        ),
                        "status": "success",
                    }
                )

            for column in numeric_columns:
                if column is primary_metric:
                    continue
                role = str(column.get("semantic_role") or "")
                if role not in {"cost", "profit", "percentage", "quantity"}:
                    continue
                title = self._metric_label(column)
                value = column.get("average") if role == "percentage" else column.get("sum")
                if not isinstance(value, (int, float)):
                    continue
                dashboard_cards.append(
                    {
                        "title": title,
                        "description": (
                            f"{'Average' if role == 'percentage' else 'Total'} {title.lower()} from the spreadsheet."
                        ),
                        "visualization": "kpi",
                        "value": value,
                        "rows": [{column.get("name"): value}],
                        "explanation": f"{title} is another measurable operating signal.",
                        "interpretation": f"{title} is currently {self._format_metric_value(value, column)}.",
                        "recommended_action": (
                            f"Compare {title.lower()} against {self._metric_label(primary_metric).lower() if primary_metric else 'the lead metric'} before making trade-offs."
                        ),
                        "status": "success",
                    }
                )

            if primary_metric and primary_date:
                trend_rows = self._aggregate_by_period(
                    rows,
                    primary_date_key,
                    primary_metric_key,
                )
                if len(trend_rows) >= 2:
                    previous = float(trend_rows[-2]["value"])
                    latest = float(trend_rows[-1]["value"])
                    delta_text = self._format_percent_delta(previous, latest)
                    tone = "positive" if latest >= previous else "warning"
                    add_insight(
                        title=f"{self._metric_label(primary_metric)} {delta_text.split()[0]} in the latest period",
                        body=(
                            f"{self._metric_label(primary_metric)} {delta_text} from "
                            f"{self._format_metric_value(previous, primary_metric)} in {trend_rows[-2]['period']} "
                            f"to {self._format_metric_value(latest, primary_metric)} in {trend_rows[-1]['period']}."
                        ),
                        tone=tone,
                        priority_score=94 if latest < previous else 88,
                        evidence={
                            "table": table.get("name"),
                            "metric": primary_metric_key,
                            "date": primary_date_key,
                            "previous": trend_rows[-2],
                            "latest": trend_rows[-1],
                        },
                    )
                    dashboard_cards.append(
                        {
                            "title": f"{self._metric_label(primary_metric)} trend",
                            "description": f"{self._metric_label(primary_metric)} by period.",
                            "visualization": "line",
                            "value": latest,
                            "rows": trend_rows[-12:],
                            "explanation": "This shows how the lead KPI has moved over time.",
                            "interpretation": f"The latest period {delta_text} versus the previous period.",
                            "recommended_action": (
                                "Investigate the segments behind the latest movement before changing spend or inventory."
                            ),
                            "status": "success",
                        }
                    )
                    if latest < previous:
                        risk = {
                            "title": f"{self._metric_label(primary_metric)} decline",
                            "body": (
                                f"The latest period is below the prior period by {abs(((latest - previous) / previous) * 100):.1f}%."
                                if previous
                                else "The latest period is below the prior period."
                            ),
                            "priority": "High",
                        }
                        risks.append(risk)
                        add_recommendation(
                            title="Investigate the latest decline",
                            body=(
                                f"Check which {self._segment_label(segment_columns[0]) if segment_columns else 'segment'} drove the drop before committing more budget."
                            ),
                            priority="High",
                            priority_score=90,
                            evidence=risk,
                        )
                    else:
                        opportunities.append(
                            {
                                "title": f"{self._metric_label(primary_metric)} momentum",
                                "body": "The latest period is outperforming the prior period.",
                                "priority": "High",
                            }
                        )

            if primary_metric and segment_columns:
                for segment in segment_columns[:3]:
                    segment_key = self._column_name(segment)
                    segment_rows = self._aggregate_by_category(
                        rows,
                        segment_key,
                        primary_metric_key,
                    )
                    if not segment_rows:
                        continue
                    top = segment_rows[0]
                    bottom = segment_rows[-1] if len(segment_rows) > 1 else None
                    segment_label = self._segment_label(segment)
                    metric_label = self._metric_label(primary_metric)
                    top_share = float(top.get("share") or 0)
                    add_insight(
                        title=f"{top['segment']} leads {metric_label.lower()}",
                        body=(
                            f"{top['segment']} contributes {self._format_metric_value(top['value'], primary_metric)} "
                            f"of {metric_label.lower()}, representing {top_share:.1f}% of the visible total by {segment_label}."
                        ),
                        tone="positive",
                        priority_score=90 if top_share >= 35 else 82,
                        evidence={
                            "table": table.get("name"),
                            "metric": primary_metric_key,
                            "segment": segment_key,
                            "top": top,
                        },
                    )
                    dashboard_cards.append(
                        {
                            "title": f"{metric_label} by {self._humanize_identifier(segment_key)}",
                            "description": f"Which {segment_label} contributes most to {metric_label.lower()}.",
                            "visualization": "bar",
                            "value": top["value"],
                            "rows": segment_rows[:10],
                            "explanation": f"This ranks {segment_label}s by their contribution to {metric_label.lower()}.",
                            "interpretation": f"{top['segment']} is the strongest visible contributor.",
                            "recommended_action": (
                                f"Protect and learn from {top['segment']}; then improve the weakest {segment_label}s."
                            ),
                            "status": "success",
                        }
                    )
                    if top_share >= 45:
                        risk = {
                            "title": f"{metric_label} concentration",
                            "body": f"{top['segment']} controls {top_share:.1f}% of {metric_label.lower()} by {segment_label}.",
                            "priority": "Medium",
                        }
                        risks.append(risk)
                        add_recommendation(
                            title=f"Reduce dependence on {top['segment']}",
                            body=(
                                f"Build a plan to grow the next strongest {segment_label}s so performance is not overly dependent on one segment."
                            ),
                            priority="High" if top_share >= 60 else "Medium",
                            priority_score=88 if top_share >= 60 else 76,
                            evidence=risk,
                        )
                    else:
                        add_recommendation(
                            title=f"Double down on {top['segment']}",
                            body=(
                                f"Use {top['segment']} as the benchmark for campaigns, inventory, or follow-up because it currently leads {metric_label.lower()}."
                            ),
                            priority="High",
                            priority_score=84,
                            evidence={"top_segment": top},
                        )
                    if bottom and float(bottom.get("value") or 0) < float(top.get("value") or 0) * 0.35:
                        opportunities.append(
                            {
                                "title": f"Improve {bottom['segment']}",
                                "body": f"{bottom['segment']} trails the leading {segment_label}, which may be a growth opportunity.",
                                "priority": "Medium",
                            }
                        )

            for status_column in [
                column for column in segment_columns if column.get("semantic_role") == "status"
            ][:2]:
                status_key = self._column_name(status_column)
                status_counts = self._count_categories(rows, status_key)
                if not status_counts:
                    continue
                top_status = status_counts[0]
                risky_statuses = [
                    item
                    for item in status_counts
                    if re.search(
                        r"cancel|churn|fail|late|lost|overdue|refund|risk|unpaid|delay",
                        str(item["segment"]).lower(),
                    )
                ]
                dashboard_cards.append(
                    {
                        "title": f"{self._humanize_identifier(status_key)} mix",
                        "description": "Operational status distribution.",
                        "visualization": "donut",
                        "value": top_status["count"],
                        "rows": status_counts[:8],
                        "explanation": "This shows where work, orders, customers, or invoices currently stand.",
                        "interpretation": f"{top_status['segment']} is the most common status at {top_status['share']:.1f}%.",
                        "recommended_action": "Prioritize the statuses that represent delays, loss, churn, or unpaid work.",
                        "status": "success",
                    }
                )
                if risky_statuses:
                    risky = risky_statuses[0]
                    risk = {
                        "title": f"{risky['segment']} needs attention",
                        "body": f"{risky['segment']} represents {risky['share']:.1f}% of records in {self._humanize_identifier(status_key)}.",
                        "priority": "High",
                    }
                    risks.append(risk)
                    add_insight(
                        title=risk["title"],
                        body=risk["body"],
                        tone="warning",
                        priority_score=92,
                        evidence=risk,
                    )
                    add_recommendation(
                        title=f"Follow up on {risky['segment']}",
                        body="Assign an owner to review the affected records and remove the operational blocker.",
                        priority="High",
                        priority_score=90,
                        evidence=risk,
                    )

            if primary_metric:
                anomalies = self._detect_numeric_anomalies(
                    rows,
                    primary_metric,
                    date_key=primary_date_key or None,
                    segment_keys=segment_keys,
                )
                if anomalies:
                    top_anomaly = anomalies[0]
                    metric_label = self._metric_label(primary_metric)
                    add_insight(
                        title=f"{metric_label} anomaly detected",
                        body=(
                            f"{top_anomaly['label']} is unusually high or low at "
                            f"{self._format_metric_value(top_anomaly['value'], primary_metric)} versus an average of "
                            f"{self._format_metric_value(top_anomaly['average'], primary_metric)}."
                        ),
                        tone="warning",
                        priority_score=86,
                        evidence={"anomaly": top_anomaly},
                    )
                    dashboard_cards.append(
                        {
                            "title": f"{metric_label} anomalies",
                            "description": "Rows that are far from the typical value.",
                            "visualization": "table",
                            "value": top_anomaly["value"],
                            "rows": anomalies,
                            "explanation": "These values are statistical outliers in the spreadsheet.",
                            "interpretation": "Outliers can signal exceptional performance, data issues, or operational risk.",
                            "recommended_action": "Review the underlying rows before treating the outlier as a trend.",
                            "status": "success",
                        }
                    )

            if not primary_metric and segment_columns:
                segment = segment_columns[0]
                counts = self._count_categories(rows, self._column_name(segment))
                if counts:
                    top = counts[0]
                    segment_label = self._segment_label(segment)
                    add_insight(
                        title=f"{top['segment']} is the largest {segment_label}",
                        body=(
                            f"{top['segment']} appears in {top['count']} records, "
                            f"{top['share']:.1f}% of the visible spreadsheet."
                        ),
                        tone="neutral",
                        priority_score=64,
                        evidence={"top_segment": top},
                    )
                    dashboard_cards.append(
                        {
                            "title": f"{self._humanize_identifier(self._column_name(segment))} distribution",
                            "description": f"Record count by {segment_label}.",
                            "visualization": "bar",
                            "value": top["count"],
                            "rows": counts[:10],
                            "explanation": f"This shows how records are distributed across {segment_label}s.",
                            "interpretation": f"{top['segment']} is the largest visible group.",
                            "recommended_action": f"Start with {top['segment']} if you need a focused operational review.",
                            "status": "success",
                        }
                    )

        insights.sort(key=lambda item: item.get("priority_score", 0), reverse=True)
        recommendations.sort(key=lambda item: item.get("priority_score", 0), reverse=True)

        if not insights:
            total_rows = sum(table.get("row_count", 0) for table in tables)
            insights.append(
                {
                    "title": "Spreadsheet evidence is available",
                    "body": (
                        f"Vayent found {total_rows:,} usable records and can answer follow-up questions from the spreadsheet contents."
                    ),
                    "tone": "neutral",
                    "priority_score": 40,
                    "evidence": {},
                }
            )
        if not recommendations:
            first_understanding = dataset_understanding[0] if dataset_understanding else {}
            primary_metric_label = (first_understanding.get("primary_metric") or {}).get("label")
            first_segment_label = (
                (first_understanding.get("segments") or [{}])[0].get("label")
                if first_understanding.get("segments")
                else None
            )
            if primary_metric_label and first_segment_label:
                recommendations.append(
                    {
                        "title": f"Review {str(primary_metric_label).lower()} by {first_segment_label}",
                        "body": (
                            f"Use the spreadsheet to compare {str(primary_metric_label).lower()} across {first_segment_label}s and prioritize the strongest or weakest group."
                        ),
                        "priority": "Medium",
                        "priority_score": 50,
                        "evidence": {},
                    }
                )
            else:
                recommendations.append(
                    {
                        "title": "Ask a follow-up business question",
                        "body": "Ask for the top performers, biggest risks, trends, anomalies, or recommendations and Vayent will use the spreadsheet rows as evidence.",
                        "priority": "Medium",
                        "priority_score": 40,
                        "evidence": {},
                    }
                )

        return {
            "tables": table_summaries,
            "dataset_understanding": dataset_understanding,
            "kpis": kpis[:12],
            "risks": risks[:8],
            "opportunities": opportunities[:8],
            "quality_checks": quality_checks,
            "suggested_questions": suggested_questions[:12],
            "insights": insights[:8],
            "recommendations": recommendations[:8],
            "dashboard_cards": dashboard_cards[:10],
        }

    def _aggregate_by_category(
        self,
        rows: list[dict[str, Any]],
        category_key: str,
        numeric_key: str,
    ) -> list[dict[str, Any]]:
        totals: dict[str, float] = {}
        counts: Counter[str] = Counter()
        for row in rows:
            segment = row.get(category_key)
            value = self._to_number(row.get(numeric_key))
            if segment in (None, "") or value is None:
                continue
            label = str(segment)
            totals[label] = totals.get(label, 0.0) + value
            counts[label] += 1

        grand_total = sum(totals.values())
        return [
            {
                "segment": key,
                "value": value,
                "count": counts[key],
                "share": (value / grand_total * 100) if grand_total else 0,
            }
            for key, value in sorted(totals.items(), key=lambda item: item[1], reverse=True)
        ]

    def process_file_bytes(
        self,
        *,
        filename: str | None,
        content_type: str | None,
        content: bytes,
    ) -> ProcessedSpreadsheet:
        file_type = self.validate_upload(
            filename=filename,
            content_type=content_type,
            content=content,
        )
        extension = self._extension_from_name(filename)
        if extension == ".csv":
            tables = [self._parse_csv(content)]
        elif extension == ".xlsx":
            tables = self._parse_xlsx(content)
        else:
            tables = self._parse_xls(content)

        raw_schema_metadata = {
            "source_type": "spreadsheet",
            "tables": [
                {
                    "name": table["name"],
                    "row_count": table["row_count"],
                    "columns": [
                        {
                            "name": column["name"],
                            "type": column["type"],
                            "semantic_role": column.get("semantic_role"),
                            "business_label": column.get("business_label"),
                            "nullable": column["nullable"],
                            "primary_key": False,
                            "foreign_key": False,
                            "foreign_key_reference": None,
                        }
                        for column in table["columns"]
                    ],
                }
                for table in tables
            ],
        }
        dataset_payload = {
            "tables": [
                {
                    "name": table["name"],
                    "row_count": table["row_count"],
                    "columns": table["columns"],
                    "rows": table["rows"],
                    "preview_rows": table.get("preview_rows", table["rows"][:5]),
                    "truncated": table.get("truncated", False),
                }
                for table in tables
            ],
        }
        analysis_metadata = self._build_analysis(tables)
        return ProcessedSpreadsheet(
            file_type=file_type,
            raw_schema_metadata=self._json_safe(raw_schema_metadata),
            dataset_payload=self._json_safe(dataset_payload),
            analysis_metadata=self._json_safe(analysis_metadata),
        )

    def _validate_url_host(self, url: str) -> None:
        parsed = urlparse(url)
        if parsed.scheme not in {"http", "https"} or not parsed.netloc:
            raise SpreadsheetValidationError("Enter a valid http or https spreadsheet URL.")
        hostname = (parsed.hostname or "").strip().lower()
        if not hostname:
            raise SpreadsheetValidationError("Spreadsheet URL is missing a hostname.")
        if hostname in {item.lower() for item in self.settings.blocked_spreadsheet_hosts}:
            raise SpreadsheetValidationError("This spreadsheet host is blocked by server policy.")
        if hostname in {"localhost", "localhost.localdomain"} and not self.settings.allow_private_spreadsheet_urls:
            raise SpreadsheetValidationError("Local spreadsheet links are disabled by server policy.")
        try:
            address = ipaddress.ip_address(hostname)
            if (
                not self.settings.allow_private_spreadsheet_urls
                and (
                    address.is_private
                    or address.is_loopback
                    or address.is_link_local
                    or address.is_multicast
                    or address.is_reserved
                    or address.is_unspecified
                )
            ):
                raise SpreadsheetValidationError("Private spreadsheet links are disabled by server policy.")
        except ValueError:
            if not self.settings.allow_private_spreadsheet_urls:
                try:
                    for result in socket.getaddrinfo(hostname, None):
                        address_text = result[4][0]
                        address = ipaddress.ip_address(address_text)
                        if (
                            address.is_private
                            or address.is_loopback
                            or address.is_link_local
                            or address.is_reserved
                            or address.is_unspecified
                        ):
                            raise SpreadsheetValidationError(
                                "Private spreadsheet links are disabled by server policy."
                            )
                except socket.gaierror as exc:
                    raise SpreadsheetValidationError("Spreadsheet link host could not be resolved.") from exc

    @staticmethod
    def normalize_spreadsheet_url(url: str) -> str:
        parsed = urlparse(url.strip())
        hostname = (parsed.hostname or "").lower()
        if hostname == "docs.google.com" and "/spreadsheets/d/" in parsed.path:
            if "/export" not in parsed.path:
                parts = parsed.path.split("/")
                try:
                    sheet_id = parts[parts.index("d") + 1]
                except (ValueError, IndexError):
                    return url.strip()
                query = parse_qs(parsed.query)
                query.setdefault("format", ["xlsx"])
                return urlunparse(
                    (
                        parsed.scheme,
                        parsed.netloc,
                        f"/spreadsheets/d/{sheet_id}/export",
                        "",
                        urlencode(query, doseq=True),
                        "",
                    )
                )
        return url.strip()

    @staticmethod
    def detect_provider(url: str) -> str:
        hostname = (urlparse(url).hostname or "").lower()
        if "docs.google.com" in hostname:
            return "google_sheets"
        if "sharepoint.com" in hostname:
            return "sharepoint"
        if "1drv.ms" in hostname or "onedrive.live.com" in hostname:
            return "onedrive"
        return "public_url"

    def _filename_from_response(self, url: str, response: httpx.Response) -> str:
        disposition = response.headers.get("content-disposition", "")
        match = re.search(r'filename\*?=(?:UTF-8\'\')?"?([^";]+)', disposition)
        if match:
            return match.group(1).strip()
        path_name = Path(urlparse(str(response.url)).path).name or Path(urlparse(url).path).name
        return path_name or "spreadsheet.xlsx"

    async def fetch_link(self, url: str) -> tuple[str, str, bytes]:
        normalized_url = self.normalize_spreadsheet_url(url)
        self._validate_url_host(normalized_url)
        try:
            async with httpx.AsyncClient(
                timeout=self.settings.spreadsheet_link_timeout_seconds,
                follow_redirects=True,
            ) as client:
                response = await client.get(normalized_url)
            response.raise_for_status()
        except httpx.HTTPError as exc:
            raise SpreadsheetValidationError(
                "Vayent could not access that spreadsheet link. Check sharing permissions or use a direct export link."
            ) from exc

        content_type = response.headers.get("content-type", "")
        normalized_content_type = content_type.split(";")[0].strip().lower()
        filename = self._filename_from_response(normalized_url, response)
        has_allowed_extension = Path(filename).suffix.lower() in self.ALLOWED_EXTENSIONS
        if (
            normalized_content_type
            and normalized_content_type not in self.SPREADSHEET_CONTENT_TYPES
            and not has_allowed_extension
        ):
            raise SpreadsheetValidationError(
                "The link did not return a compatible spreadsheet file."
            )
        return filename, content_type, response.content

    async def create_upload_source(
        self,
        *,
        user_id: str,
        name: str,
        filename: str | None,
        content_type: str | None,
        content: bytes,
        db: AsyncSession,
    ) -> SpreadsheetSource:
        processed = self.process_file_bytes(
            filename=filename,
            content_type=content_type,
            content=content,
        )
        source = SpreadsheetSource(
            id=str(uuid.uuid4()),
            user_id=user_id,
            name=self._clean_name(name or filename or "Spreadsheet"),
            source_kind=SpreadsheetSourceKind.UPLOAD,
            file_type=processed.file_type,
            original_filename=filename,
            source_provider="upload",
            status="connected",
            status_message="Spreadsheet uploaded and profiled.",
            raw_schema_metadata=processed.raw_schema_metadata,
            dataset_payload=processed.dataset_payload,
            analysis_metadata=processed.analysis_metadata,
            is_active=True,
            last_synced_at=datetime.utcnow(),
        )
        db.add(source)
        await db.commit()
        await db.refresh(source)
        return source

    async def create_link_source(
        self,
        *,
        user_id: str,
        name: str,
        url: str,
        db: AsyncSession,
    ) -> SpreadsheetSource:
        normalized_url = self.normalize_spreadsheet_url(url)
        filename, content_type, content = await self.fetch_link(normalized_url)
        processed = self.process_file_bytes(
            filename=filename,
            content_type=content_type,
            content=content,
        )
        source = SpreadsheetSource(
            id=str(uuid.uuid4()),
            user_id=user_id,
            name=self._clean_name(name or filename or "Spreadsheet link"),
            source_kind=SpreadsheetSourceKind.LINK,
            file_type=processed.file_type,
            original_filename=filename,
            source_url=normalized_url,
            source_provider=self.detect_provider(normalized_url),
            status="connected",
            status_message="Spreadsheet link connected and profiled.",
            raw_schema_metadata=processed.raw_schema_metadata,
            dataset_payload=processed.dataset_payload,
            analysis_metadata=processed.analysis_metadata,
            is_active=True,
            last_synced_at=datetime.utcnow(),
        )
        db.add(source)
        await db.commit()
        await db.refresh(source)
        return source

    async def get_source(self, source_id: str, db: AsyncSession) -> SpreadsheetSource | None:
        result = await db.execute(
            select(SpreadsheetSource).where(SpreadsheetSource.id == source_id)
        )
        return result.scalar_one_or_none()

    async def get_user_sources(self, user_id: str, db: AsyncSession) -> list[SpreadsheetSource]:
        result = await db.execute(
            select(SpreadsheetSource)
            .where(
                SpreadsheetSource.user_id == user_id,
                SpreadsheetSource.is_active == True,
            )
            .order_by(SpreadsheetSource.created_at.desc())
        )
        return list(result.scalars().all())

    async def rename_source(
        self,
        *,
        source_id: str,
        user_id: str,
        name: str,
        db: AsyncSession,
    ) -> SpreadsheetSource | None:
        source = await self.get_source(source_id, db)
        if not source or source.user_id != user_id or not source.is_active:
            return None
        source.name = self._clean_name(name)
        source.updated_at = datetime.utcnow()
        await db.commit()
        await db.refresh(source)
        return source

    async def disconnect_source(self, *, source_id: str, user_id: str, db: AsyncSession) -> bool:
        source = await self.get_source(source_id, db)
        if not source or source.user_id != user_id:
            return False
        source.is_active = False
        source.status = "disconnected"
        source.updated_at = datetime.utcnow()
        await db.commit()
        return True

    async def sync_source(
        self,
        *,
        source_id: str,
        user_id: str,
        db: AsyncSession,
    ) -> SpreadsheetSource:
        source = await self.get_source(source_id, db)
        if not source or source.user_id != user_id or not source.is_active:
            raise ValueError("Spreadsheet source not found")

        try:
            source_kind = (
                source.source_kind.value
                if hasattr(source.source_kind, "value")
                else str(source.source_kind)
            )
            if source_kind == SpreadsheetSourceKind.LINK.value:
                if not source.source_url:
                    raise SpreadsheetValidationError("Spreadsheet link is missing.")
                filename, content_type, content = await self.fetch_link(source.source_url)
                processed = self.process_file_bytes(
                    filename=filename,
                    content_type=content_type,
                    content=content,
                )
                source.original_filename = filename
                source.file_type = processed.file_type
                source.raw_schema_metadata = processed.raw_schema_metadata
                source.dataset_payload = processed.dataset_payload
                source.analysis_metadata = processed.analysis_metadata
                source.status_message = "Spreadsheet link synced and reprocessed."
            else:
                tables = (source.dataset_payload or {}).get("tables", [])
                source.analysis_metadata = self._json_safe(self._build_analysis(tables))
                source.status_message = "Uploaded spreadsheet metadata refreshed."

            source.status = "connected"
            source.last_synced_at = datetime.utcnow()
            source.updated_at = datetime.utcnow()
            await db.commit()
            await db.refresh(source)
            return source
        except Exception as exc:
            source.status = "error"
            source.status_message = str(exc)
            source.updated_at = datetime.utcnow()
            await db.commit()
            raise

    def format_source_for_ai(self, source: SpreadsheetSource) -> str:
        analysis = source.analysis_metadata or {}
        parts = [
            f"Spreadsheet Source: {source.name}",
            f"Connection method: {source.source_kind.value if hasattr(source.source_kind, 'value') else source.source_kind}",
            f"File type: {source.file_type}",
            "Use spreadsheet columns and computed evidence as data fields. Explain insights in business language.",
        ]
        understanding = analysis.get("dataset_understanding") or []
        if understanding:
            parts.append("")
            parts.append("Dataset understanding:")
            for item in understanding[:3]:
                metric = item.get("primary_metric") or {}
                segments = item.get("segments") or []
                segment_labels = [
                    str(segment.get("label") or segment.get("column"))
                    for segment in segments[:4]
                    if segment.get("label") or segment.get("column")
                ]
                parts.append(
                    "  - "
                    f"{item.get('table')}: lead metric={metric.get('label') or metric.get('column') or 'none'}, "
                    f"date={item.get('date_column') or 'none'}, "
                    f"segments={', '.join(segment_labels) or 'none'}"
                )
        if analysis.get("insights"):
            parts.append("")
            parts.append("Computed business insights:")
            for insight in analysis.get("insights", [])[:5]:
                parts.append(f"  - {insight.get('title')}: {insight.get('body')}")
        if analysis.get("recommendations"):
            parts.append("")
            parts.append("Recommended actions:")
            for recommendation in analysis.get("recommendations", [])[:4]:
                parts.append(
                    f"  - {recommendation.get('priority', 'Medium')}: {recommendation.get('body')}"
                )
        if analysis.get("quality_checks"):
            parts.append("")
            parts.append("Quality checks:")
            for check in analysis.get("quality_checks", [])[:3]:
                parts.append(
                    f"  - {check.get('table')}: missing fields={len(check.get('missing_values') or [])}, "
                    f"duplicates={check.get('duplicate_records', 0)}, invalid date columns={len(check.get('invalid_dates') or [])}"
                )
        if analysis.get("suggested_questions"):
            parts.append("")
            parts.append("Suggested questions:")
            for question in analysis.get("suggested_questions", [])[:6]:
                parts.append(f"  - {question}")
        for table in (source.dataset_payload or {}).get("tables", []):
            parts.append("")
            parts.append(f"Sheet: {table.get('name')}")
            parts.append(f"Rows: {table.get('row_count', 0)}")
            parts.append("Columns:")
            for column in table.get("columns", []):
                parts.append(
                    f"  - {column.get('name')} ({column.get('type')}); role: {column.get('semantic_role')}; label: {column.get('business_label') or column.get('label')}"
                )
            rows = (table.get("preview_rows") or table.get("rows", []))[:5]
            if rows:
                parts.append("Sample rows:")
                for row in rows:
                    parts.append(f"  - {row}")
        return "\n".join(parts)

    @staticmethod
    def _prompt_token_set(prompt: str) -> set[str]:
        stopwords = {
            "about",
            "also",
            "and",
            "any",
            "are",
            "biggest",
            "can",
            "for",
            "from",
            "give",
            "insight",
            "insights",
            "into",
            "me",
            "my",
            "of",
            "on",
            "show",
            "spreadsheet",
            "tell",
            "the",
            "this",
            "what",
            "which",
            "with",
        }
        tokens = re.findall(r"[a-z0-9]+", prompt.lower())
        normalized_tokens: set[str] = set()
        for token in tokens:
            if len(token) < 3 or token in stopwords:
                continue
            normalized_tokens.add(token)
            if token.endswith("ies") and len(token) > 4:
                normalized_tokens.add(f"{token[:-3]}y")
            elif token.endswith("s") and len(token) > 3:
                normalized_tokens.add(token[:-1])
        return normalized_tokens

    @staticmethod
    def _pluralize_label(label: str, count: int | None = None) -> str:
        cleaned = label.strip()
        if not cleaned:
            return "records"
        if count == 1:
            return cleaned
        if cleaned.endswith("y") and not cleaned.endswith(("ay", "ey", "iy", "oy", "uy")):
            return f"{cleaned[:-1]}ies"
        if cleaned.endswith("s"):
            return cleaned
        return f"{cleaned}s"

    def _score_prompt_match(self, prompt_tokens: set[str], *values: Any) -> int:
        if not prompt_tokens:
            return 0
        haystack = self._normalized_identifier(
            " ".join(str(value) for value in values if value not in (None, ""))
        )
        return sum(1 for token in prompt_tokens if token in haystack)

    def _select_prompt_items(
        self,
        items: list[dict[str, Any]],
        *,
        prompt: str,
        limit: int,
    ) -> list[dict[str, Any]]:
        prompt_tokens = self._prompt_token_set(prompt)
        prompt_text = prompt.lower()
        scored = []
        for index, item in enumerate(items):
            score = self._score_prompt_match(
                prompt_tokens,
                item.get("title"),
                item.get("body"),
                item.get("description"),
                item.get("explanation"),
                item.get("interpretation"),
                item.get("recommended_action"),
                item.get("visualization"),
            )
            if "risk" in prompt_text and str(item.get("tone", "")).lower() == "warning":
                score += 4
            if "trend" in prompt_text and item.get("visualization") in {"line", "area"}:
                score += 4
            if any(word in prompt_text for word in ("top", "best", "highest", "leading")) and item.get("visualization") == "bar":
                score += 4
            if any(word in prompt_text for word in ("anomaly", "outlier", "unusual")) and "anomal" in str(item.get("title", "")).lower():
                score += 5
            scored.append((score, -index, item))
        ranked = [item for score, _, item in sorted(scored, reverse=True) if score > 0]
        if not ranked:
            ranked = items
        return ranked[:limit]

    def _matching_spreadsheet_rows(
        self,
        source: SpreadsheetSource,
        *,
        prompt: str,
        limit: int = 12,
    ) -> list[dict[str, Any]]:
        prompt_text = self._normalized_identifier(prompt)
        prompt_tokens = self._prompt_token_set(prompt)
        if not prompt_tokens:
            return []

        matches: list[dict[str, Any]] = []
        for table in (source.dataset_payload or {}).get("tables", []):
            table_name = table.get("name") or "Sheet"
            for row in table.get("rows", []):
                values = [value for value in row.values() if value not in (None, "")]
                row_text = self._normalized_identifier(" ".join(str(value) for value in values))
                exact_value_match = any(
                    len(str(value)) >= 3
                    and self._normalized_identifier(str(value)) in prompt_text
                    for value in values
                )
                token_overlap = sum(1 for token in prompt_tokens if token in row_text)
                if exact_value_match or token_overlap >= 2:
                    enriched = {"sheet": table_name, **self._json_safe(row)}
                    matches.append(enriched)
                if len(matches) >= limit:
                    return matches
        return matches

    @staticmethod
    def _rows_from_cards(cards: list[dict[str, Any]], *, limit: int = 20) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for card in cards:
            for row in card.get("rows", []) or []:
                if not isinstance(row, dict):
                    continue
                enriched = {"analysis": card.get("title"), **row}
                rows.append(enriched)
                if len(rows) >= limit:
                    return rows
        return rows

    @staticmethod
    def _prompt_limit(prompt: str, *, default: int = 10, maximum: int = 50) -> int:
        match = re.search(r"\b(?:list|show|first|latest|recent|top)\s+(\d{1,3})\b", prompt.lower())
        if not match:
            match = re.search(r"\b(\d{1,3})\b", prompt.lower())
        if not match:
            return default
        return max(1, min(int(match.group(1)), maximum))

    def _source_table_contexts(self, source: SpreadsheetSource) -> list[dict[str, Any]]:
        contexts = []
        for table in (source.dataset_payload or {}).get("tables", []):
            columns = table.get("columns", []) or []
            rows = table.get("rows", []) or []
            contexts.append(
                {
                    "table": table,
                    "columns": columns,
                    "rows": rows,
                    "entity": self._select_entity_column(columns),
                    "date": self._select_primary_date(columns),
                    "segments": self._select_segment_columns(columns),
                }
            )
        return contexts

    def _enrich_spreadsheet_row(
        self,
        *,
        table_name: str,
        row: dict[str, Any],
        columns: list[dict[str, Any]],
        entity_column: dict[str, Any] | None = None,
        date_column: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        enriched = {"sheet": table_name}
        if entity_column:
            entity_key = self._column_name(entity_column)
            if row.get(entity_key) not in (None, ""):
                enriched[self._segment_label(entity_column).replace(" ", "_")] = row.get(entity_key)
        if date_column:
            date_key = self._column_name(date_column)
            if row.get(date_key) not in (None, ""):
                enriched["date"] = row.get(date_key)
        for column in columns:
            key = self._column_name(column)
            if key in enriched or row.get(key) in (None, ""):
                continue
            enriched[key] = row.get(key)
            if len(enriched) >= 8:
                break
        return self._json_safe(enriched)

    def _column_matches_prompt(
        self,
        column: dict[str, Any],
        prompt_tokens: set[str],
        prompt_text: str,
    ) -> bool:
        labels = [
            column.get("name"),
            column.get("label"),
            column.get("business_label"),
            column.get("semantic_role"),
        ]
        normalized = self._normalized_identifier(" ".join(str(item) for item in labels if item))
        if normalized and normalized in prompt_text:
            return True
        return any(token in normalized for token in prompt_tokens)

    def _select_count_column(
        self,
        *,
        columns: list[dict[str, Any]],
        prompt_tokens: set[str],
        prompt_text: str,
    ) -> dict[str, Any] | None:
        categorical_columns = [
            column for column in columns if self._is_segment_column(column)
        ]
        for column in categorical_columns:
            if any(
                self._normalized_identifier(str(category.get("value"))) in prompt_text
                for category in column.get("categories", []) or []
                if category.get("value") not in (None, "")
            ):
                return column

        matched = [
            column
            for column in categorical_columns
            if self._column_matches_prompt(column, prompt_tokens, prompt_text)
        ]
        if matched:
            return matched[0]
        status = next(
            (column for column in columns if column.get("semantic_role") == "status"),
            None,
        )
        geo = next(
            (column for column in columns if column.get("semantic_role") == "geography"),
            None,
        )
        if geo and any(token in prompt_text for token in {"state", "region", "market", "country", "city"}):
            return geo
        return status

    def _matching_category_value(
        self,
        *,
        column: dict[str, Any],
        prompt_text: str,
    ) -> str | None:
        for category in column.get("categories", []) or []:
            value = str(category.get("value") or "").strip()
            if not value:
                continue
            normalized_value = self._normalized_identifier(value)
            if normalized_value and normalized_value in prompt_text:
                return value
        return None

    def _build_count_query_result(
        self,
        *,
        source: SpreadsheetSource,
        prompt: str,
    ) -> dict[str, Any] | None:
        prompt_text = self._normalized_identifier(prompt)
        prompt_tokens = self._prompt_token_set(prompt)
        count_words = {"count", "many", "number", "total", "registered", "registrations"}
        group_words = {"each", "by", "breakdown", "distribution", "highest", "most"}
        if not (count_words & prompt_tokens or "how many" in prompt.lower()):
            return None

        result_rows: list[dict[str, Any]] = []
        response_parts: list[str] = []
        for context in self._source_table_contexts(source):
            table = context["table"]
            rows = context["rows"]
            columns = context["columns"]
            entity = context["entity"]
            count_column = self._select_count_column(
                columns=columns,
                prompt_tokens=prompt_tokens,
                prompt_text=prompt_text,
            )
            entity_label = self._segment_label(entity).lower() if entity else "records"
            plural_entity_label = self._pluralize_label(entity_label)

            if count_column:
                category_value = self._matching_category_value(
                    column=count_column,
                    prompt_text=prompt_text,
                )
                column_key = self._column_name(count_column)
                if category_value:
                    matched_count = sum(
                        1
                        for row in rows
                        if self._normalized_identifier(str(row.get(column_key, "")))
                        == self._normalized_identifier(category_value)
                    )
                    label = f"{category_value} {plural_entity_label}"
                    result_rows.append(
                        {
                            "sheet": table.get("name"),
                            "metric": label,
                            "value": matched_count,
                            "column": column_key,
                            "category": category_value,
                        }
                    )
                    response_parts.append(f"{matched_count:,} {label} found")
                    continue

                if group_words & prompt_tokens or "highest" in prompt_tokens or "most" in prompt_tokens:
                    counts = self._count_categories(rows, column_key)
                    for item in counts[:12]:
                        result_rows.append(
                            {
                                "sheet": table.get("name"),
                                "segment": item["segment"],
                                "count": item["count"],
                                "share": item["share"],
                                "column": column_key,
                            }
                        )
                    if counts:
                        top = counts[0]
                        response_parts.append(
                            f"{top['segment']} has the highest count with {top['count']:,} records"
                        )
                    continue

            result_rows.append(
                {
                    "sheet": table.get("name"),
                    "metric": f"registered {plural_entity_label}",
                    "value": len(rows),
                }
            )
            response_parts.append(f"{len(rows):,} registered {self._pluralize_label(entity_label, len(rows))}")

        if not result_rows:
            return None
        return {
            "kind": "count",
            "rows": result_rows,
            "response": "; ".join(response_parts) + ".",
        }

    def _build_list_query_result(
        self,
        *,
        source: SpreadsheetSource,
        prompt: str,
    ) -> dict[str, Any] | None:
        prompt_lower = prompt.lower()
        prompt_tokens = self._prompt_token_set(prompt)
        explicit_list_terms = {
            "list",
            "first",
            "record",
            "row",
            "example",
            "recent",
            "latest",
            "newest",
            "recently",
        }
        count_terms = {"count", "many", "number", "total"}
        if (count_terms & prompt_tokens) and not (explicit_list_terms & prompt_tokens):
            return None
        list_requested = any(
            phrase in prompt_lower
            for phrase in (
                "list",
                "show",
                "first",
                "records",
                "rows",
                "examples",
                "recent",
                "latest",
                "newest",
            )
        )
        if not list_requested:
            return None

        limit = self._prompt_limit(prompt, default=10, maximum=50)
        want_recent = any(word in prompt_lower for word in ("recent", "latest", "newest"))
        result_rows: list[dict[str, Any]] = []
        response_parts: list[str] = []

        for context in self._source_table_contexts(source):
            table = context["table"]
            rows = list(context["rows"])
            columns = context["columns"]
            entity = context["entity"]
            date_column = context["date"]
            if want_recent and date_column:
                date_key = self._column_name(date_column)
                rows = sorted(
                    rows,
                    key=lambda row: self._to_datetime(row.get(date_key)) or datetime.min,
                    reverse=True,
                )
            for row in rows[: max(0, limit - len(result_rows))]:
                result_rows.append(
                    self._enrich_spreadsheet_row(
                        table_name=table.get("name") or "Sheet",
                        row=row,
                        columns=columns,
                        entity_column=entity,
                        date_column=date_column,
                    )
                )
            if result_rows:
                entity_label = self._segment_label(entity).lower() if entity else "records"
                modifier = "most recent " if want_recent else ""
                response_parts.append(
                    f"Showing {len(result_rows):,} {modifier}{self._pluralize_label(entity_label, len(result_rows))}."
                )
                break

        if not result_rows:
            return None
        return {
            "kind": "list",
            "rows": result_rows,
            "response": " ".join(response_parts),
        }

    def _build_search_query_result(
        self,
        *,
        source: SpreadsheetSource,
        prompt: str,
    ) -> dict[str, Any] | None:
        matched_rows = self._matching_spreadsheet_rows(source, prompt=prompt, limit=20)
        if not matched_rows:
            return None
        return {
            "kind": "search",
            "rows": matched_rows,
            "response": f"Found {len(matched_rows):,} matching spreadsheet rows.",
        }

    def _build_duplicate_query_result(
        self,
        *,
        source: SpreadsheetSource,
        prompt: str,
    ) -> dict[str, Any] | None:
        if "duplicate" not in prompt.lower() and "duplicates" not in prompt.lower():
            return None

        duplicate_rows: list[dict[str, Any]] = []
        for context in self._source_table_contexts(source):
            table = context["table"]
            rows = context["rows"]
            columns = context["columns"]
            entity = context["entity"]
            date_column = context["date"]
            fingerprints: dict[str, list[dict[str, Any]]] = {}
            for row in rows:
                fingerprint = json_dumps(
                    {
                        key: row.get(key)
                        for key in sorted(row.keys())
                        if row.get(key) not in (None, "")
                    }
                )
                fingerprints.setdefault(fingerprint, []).append(row)
            for grouped_rows in fingerprints.values():
                if len(grouped_rows) <= 1:
                    continue
                for row in grouped_rows[:5]:
                    duplicate_rows.append(
                        self._enrich_spreadsheet_row(
                            table_name=table.get("name") or "Sheet",
                            row=row,
                            columns=columns,
                            entity_column=entity,
                            date_column=date_column,
                        )
                    )
                if len(duplicate_rows) >= 20:
                    break
            if len(duplicate_rows) >= 20:
                break

        if not duplicate_rows:
            return {
                "kind": "duplicates",
                "rows": [{"metric": "duplicate_records", "value": 0}],
                "response": "I did not find duplicate records in the spreadsheet rows Vayent analyzed.",
            }
        return {
            "kind": "duplicates",
            "rows": duplicate_rows,
            "response": f"Found {len(duplicate_rows):,} duplicate record rows in the spreadsheet evidence.",
        }

    def _build_spreadsheet_query_result(
        self,
        *,
        source: SpreadsheetSource,
        prompt: str,
    ) -> dict[str, Any] | None:
        for builder in (
            self._build_duplicate_query_result,
            self._build_list_query_result,
            self._build_count_query_result,
            self._build_search_query_result,
        ):
            result = builder(source=source, prompt=prompt)
            if result:
                return result
        return None

    def _compose_spreadsheet_response(
        self,
        *,
        source: SpreadsheetSource,
        user_prompt: str,
        selected_insights: list[dict[str, Any]],
        selected_cards: list[dict[str, Any]],
        matched_rows: list[dict[str, Any]],
        direct_response: str | None = None,
    ) -> str:
        analysis = source.analysis_metadata or {}
        insights = selected_insights or (analysis.get("insights") or [])
        recommendations = analysis.get("recommendations") or []
        risks = analysis.get("risks") or []
        opportunities = analysis.get("opportunities") or []
        total_rows = sum(
            table.get("row_count", 0)
            for table in (source.dataset_payload or {}).get("tables", [])
        )

        response_parts = [
            f"I analyzed {source.name} using {total_rows:,} spreadsheet rows."
        ]
        if direct_response:
            response_parts.append(direct_response)
        if matched_rows:
            response_parts.append(
                f"I found {len(matched_rows)} row-level matches for the terms in your question."
            )
        if insights:
            response_parts.append(
                f"Most important finding: {insights[0].get('body') or insights[0].get('title')}."
            )

        risk = risks[0] if risks else next(
            (
                insight
                for insight in insights
                if str(insight.get("tone", "")).lower() == "warning"
            ),
            None,
        )
        if risk:
            response_parts.append(
                f"Biggest risk: {risk.get('body') or risk.get('title')}."
            )

        opportunity = opportunities[0] if opportunities else next(
            (
                insight
                for insight in insights[1:]
                if str(insight.get("tone", "")).lower() == "positive"
            ),
            None,
        )
        if opportunity:
            response_parts.append(
                f"Biggest opportunity: {opportunity.get('body') or opportunity.get('title')}."
            )

        if recommendations:
            response_parts.append(
                f"Recommended action: {recommendations[0].get('body') or recommendations[0].get('title')}."
            )
        elif selected_cards:
            action = selected_cards[0].get("recommended_action")
            if action:
                response_parts.append(f"Recommended action: {action}.")

        return " ".join(part.strip() for part in response_parts if part)

    def build_workspace_summary(
        self,
        *,
        source: SpreadsheetSource,
        user_prompt: str,
    ) -> dict[str, Any]:
        analysis = source.analysis_metadata or {}
        insights = analysis.get("insights") or []
        cards = analysis.get("dashboard_cards") or []
        selected_insights = self._select_prompt_items(
            insights,
            prompt=user_prompt,
            limit=4,
        )
        selected_cards = self._select_prompt_items(
            cards,
            prompt=user_prompt,
            limit=4,
        )
        query_result = self._build_spreadsheet_query_result(
            source=source,
            prompt=user_prompt,
        )
        matched_rows = (
            self._matching_spreadsheet_rows(
                source,
                prompt=user_prompt,
                limit=12,
            )
            if not query_result
            else []
        )
        # Collect a limited set of raw source rows as evidence so responses
        # can remain traceable to the underlying spreadsheet data.
        source_rows: list[dict[str, Any]] = []
        max_source_rows = 50
        for table in (source.dataset_payload or {}).get("tables", []):
            for row in (table.get("rows") or [])[: max(0, max_source_rows - len(source_rows))]:
                enriched = {"sheet": table.get("name"), **self._json_safe(row)}
                source_rows.append(enriched)
            if len(source_rows) >= max_source_rows:
                break

        # Decide which rows to present as the primary evidence for AI responses.
        # Prefer explicit query results, then matched rows, then live sampled
        # source rows; fall back to card-derived rows if nothing else is found.
        if query_result:
            rows = query_result.get("rows", [])
        elif matched_rows:
            rows = matched_rows
        else:
            sample_limit = self._prompt_limit(user_prompt, default=10, maximum=20)
            rows = source_rows[:sample_limit] or self._rows_from_cards(selected_cards, limit=20)

        # If still empty, include preview rows from the dataset payload as a last resort.
        if not rows:
            for table in (source.dataset_payload or {}).get("tables", []):
                for row in (table.get("preview_rows") or table.get("rows", []))[:5]:
                    rows.append({"sheet": table.get("name"), **self._json_safe(row)})
                if rows:
                    break
        response = self._compose_spreadsheet_response(
            source=source,
            user_prompt=user_prompt,
            selected_insights=selected_insights,
            selected_cards=selected_cards,
            matched_rows=matched_rows,
            direct_response=query_result.get("response") if query_result else None,
        )
        return {
            "source_id": source.id,
            "source_type": "spreadsheet",
            "connection_id": source.id,
            "connection_name": source.name,
            "database_name": source.original_filename or source.name,
            "sql": "SPREADSHEET_ANALYSIS",
            "row_count": len(rows),
            "truncated": len(rows) >= 20,
            "rows": rows,
            "source_rows": source_rows,
            "error": None,
            "response": response,
            "prompt": user_prompt,
            "analysis_kind": query_result.get("kind") if query_result else "insight",
            "insights": selected_insights,
            "recommendations": analysis.get("recommendations") or [],
            "risks": analysis.get("risks") or [],
            "opportunities": analysis.get("opportunities") or [],
        }

    def build_dashboard_payload(
        self,
        *,
        source: SpreadsheetSource,
        prompt: str,
    ) -> dict[str, Any]:
        analysis = source.analysis_metadata or {}
        cards = analysis.get("dashboard_cards") or []
        if not cards:
            cards = [
                {
                    "title": "Spreadsheet records",
                    "description": "How much evidence is available in this source.",
                    "visualization": "kpi",
                    "value": sum(
                        table.get("row_count", 0)
                        for table in (source.dataset_payload or {}).get("tables", [])
                    ),
                    "rows": [],
                    "explanation": "The spreadsheet is connected and available for row-level questions.",
                    "interpretation": "Use chat to ask for top performers, risks, opportunities, or anomalies in this source.",
                    "recommended_action": "Ask Vayent which records, customers, products, or regions need attention first.",
                    "status": "success",
                }
            ]
        title = f"{source.name} business dashboard"
        return {
            "title": title,
            "summary": "Business-focused dashboard generated from spreadsheet patterns.",
            "payload": {
                "description": f"Dashboard generated from {source.name}.",
                "source_ids": [source.id],
                "source_type": "spreadsheet",
                "source_name": source.name,
                "cards": cards,
                "insights": analysis.get("insights") or [],
                "recommendations": analysis.get("recommendations") or [],
                "kpis": analysis.get("kpis") or [],
                "risks": analysis.get("risks") or [],
                "opportunities": analysis.get("opportunities") or [],
                "quality_checks": analysis.get("quality_checks") or [],
                "suggested_questions": analysis.get("suggested_questions") or [],
                "dataset_understanding": analysis.get("dataset_understanding") or [],
                "prompt": prompt,
            },
        }


def json_dumps(value: Any) -> str:
    import json

    return json.dumps(value, default=str)


def json_loads(value: str) -> Any:
    import json

    return json.loads(value)


spreadsheet_service = SpreadsheetService()
